import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side


def read_epicycle_data(filename):
    # 从文本文件中读取数据
    data = []
    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            parts = line.strip().split(', ')
            freq = float(parts[0].split(': ')[1])
            radius = float(parts[1].split(': ')[1])
            phase = float(parts[2].split(': ')[1])
            data.append({'频率': freq, '半径': radius, '相位': phase})
    return data


def calculate_coordinates(data):
    # 计算每个周转圆的初始位置、圆心坐标和相对频率
    coordinates = []
    x, y = 0, 0  # 初始化圆心坐标
    previous_frequency = 0  # 用于计算相对频率

    for i, circle in enumerate(data):
        freq = circle['频率']
        radius = circle['半径']
        phase = circle['相位']
        phase_deg = math.degrees(phase)  # 转换相位到度数制

        # 计算相对频率
        relative_freq = freq - previous_frequency
        previous_frequency = freq  # 更新上一频率

        # 计算当前圆的相对坐标
        cx = x + radius * math.cos(phase)  # 当前圆心的x坐标
        cy = y + radius * math.sin(phase)  # 当前圆心的y坐标

        coordinates.append({
            '序号': i + 1,
            '绝对转速': freq,
            '相对转速': relative_freq,
            '半径 (单位)': radius,
            '相位 (度)': phase_deg,
            '圆心X坐标 (单位)': cx,
            '圆心Y坐标 (单位)': cy
        })

        # 更新下一个圆的起点为当前圆的圆心
        x, y = cx, cy

    return coordinates


def save_to_excel(data, output_filename):
    # 将数据保存到Excel文件
    df = pd.DataFrame(data)
    df.to_excel(output_filename, index=False)

    # 使用openpyxl打开生成的Excel文件
    wb = load_workbook(output_filename)
    ws = wb.active

    # 设置单元格样式
    alignment = Alignment(horizontal='center', vertical='center')
    fill = PatternFill(start_color='FFCCCCFF', end_color='FFCCCCFF', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 调整列宽和应用样式
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # 获取列字母
        for cell in column:
            cell.alignment = alignment
            cell.fill = fill
            cell.border = thin_border
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 5  # 调整列宽

    wb.save(output_filename)
    print(f"数据已成功保存到 {output_filename}.")


if __name__ == "__main__":
    input_filename = 'epicycles.txt'  # 输入的文本文件名
    output_filename = '机械师专用.xlsx'  # 输出的Excel文件名

    # 读取数据并计算坐标
    epicycle_data = read_epicycle_data(input_filename)
    coordinate_data = calculate_coordinates(epicycle_data)

    # 保存结果到Excel文件并进行格式化
    save_to_excel(coordinate_data, output_filename)
